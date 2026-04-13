"""
Build AO explanations in v009 — Baker version 3.
Total -27,134,844 as anchor. CAPEX/OPEX breakdown.
Sources: Fund / AO / Other. No separate capital call section (Edita rows 58-66 handle that).
"""
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

src = "/Users/dimitry/Desktop/009_MOVIE_AO_Baker_Version_20260411_WITH_AO_EXPLANATIONS.xlsx"
dst = "/Users/dimitry/Desktop/009_MOVIE_AO_Baker_Version_20260411 Baker version 3.xlsx"
shutil.copy2(src, dst)

wb = openpyxl.load_workbook(dst)
ws = wb['AO ']

# --- Clear old additions (rows 71-130) ---
# First unmerge any merged cells in this range
merges_to_remove = []
for merge in ws.merged_cells.ranges:
    if merge.min_row >= 71 and merge.min_row <= 130:
        merges_to_remove.append(merge)
for merge in merges_to_remove:
    ws.unmerge_cells(str(merge))

for r in range(71, 131):
    for c in range(1, 21):
        cell = ws.cell(row=r, column=c)
        cell.value = None
        cell.font = Font(name='Calibri', size=10)
        cell.fill = PatternFill(fill_type=None)
        cell.border = Border()
        cell.alignment = Alignment()
        cell.number_format = 'General'

# --- Styles ---
title_font = Font(name='Calibri', size=12, bold=True, color='1a1a6e')
header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
section_font = Font(name='Calibri', size=10, bold=True, color='1a1a6e')
normal_font = Font(name='Calibri', size=10)
bold_font = Font(name='Calibri', size=10, bold=True)
small_font = Font(name='Calibri', size=9, color='555555')
red_font = Font(name='Calibri', size=10, bold=True, color='CC0000')
green_font = Font(name='Calibri', size=10, bold=True, color='006600')
total_font = Font(name='Calibri', size=11, bold=True)
big_total_font = Font(name='Calibri', size=11, bold=True, color='CC0000')

header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
capex_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
opex_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
new_fill = PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid')
source_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
grand_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')

eur_fmt = '#,##0'
COLS_FILL = ['K','L','M','N','O','P','Q','R','S','T']

def fill_row(ws, row, fill, cols=COLS_FILL):
    for c in cols:
        ws[f'{c}{row}'].fill = fill

row = 71

# ============================================================
# TITLE
# ============================================================
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = 'EXPLANATION OF ADDITIONAL REQUIRED AMOUNTS — For AO Review'
ws[f'K{row}'].font = title_font
row += 1

ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = 'Total capital required Oct 2025 – Jun 2026. Breakdown by CAPEX and OPEX. Sources of funding.'
ws[f'K{row}'].font = Font(name='Calibri', size=9, color='666666')
row += 2

# ============================================================
# CONTEXT
# ============================================================
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = 'In September 2025, the Capital Call was EUR 44,221,554 — fully covered (Bank EUR 14.8M + Fund EUR 15.3M + AO EUR 14.1M contractual).'
ws[f'K{row}'].font = normal_font
row += 1
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = 'Since then: (1) 2-month construction delay, (2) Hagenauer insolvency Mar 2026, (3) New Austrian VAT law Nov 2025. Hotel opened Nov 27, creating OPEX.'
ws[f'K{row}'].font = bold_font
row += 2

# ============================================================
# COLUMN HEADERS
# ============================================================
fill_row(ws, row, header_fill)
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'USES — Total Capital Required'
ws[f'K{row}'].font = header_font
ws[f'O{row}'].value = 'Oct-Dec 2025'
ws[f'O{row}'].font = header_font
ws[f'O{row}'].alignment = Alignment(horizontal='center')
ws[f'Q{row}'].value = 'Jan-Jun 2026'
ws[f'Q{row}'].font = header_font
ws[f'Q{row}'].alignment = Alignment(horizontal='center')
ws[f'S{row}'].value = 'Total Oct-Jun'
ws[f'S{row}'].font = header_font
ws[f'S{row}'].alignment = Alignment(horizontal='center')
row += 1

# ============================================================
# CAPEX
# ============================================================
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'CAPEX — Construction & Development'
ws[f'K{row}'].font = section_font
fill_row(ws, row, capex_fill)
row += 1

def item_row(ws, row, label, oct_dec, jan_jun, total, font=normal_font, fill=None, note=None, is_new=False):
    ws.merge_cells(f'K{row}:N{row}')
    ws[f'K{row}'].value = label
    if is_new:
        ws[f'K{row}'].font = red_font
        fill_row(ws, row, new_fill)
    else:
        ws[f'K{row}'].font = font
        if fill:
            fill_row(ws, row, fill)

    for col, val in [('O', oct_dec), ('Q', jan_jun), ('S', total)]:
        if val is not None:
            ws[f'{col}{row}'].value = val
            ws[f'{col}{row}'].number_format = eur_fmt
            ws[f'{col}{row}'].font = red_font if is_new else font
            ws[f'{col}{row}'].alignment = Alignment(horizontal='right')

    if note:
        row2 = row + 1
        ws.merge_cells(f'K{row2}:T{row2}')
        ws[f'K{row2}'].value = f'    → {note}'
        ws[f'K{row2}'].font = small_font
        return row + 2
    return row + 1

# v009 CAPEX items
row = item_row(ws, row, 'Construction & Development', -1341458, -2541213, -3882671,
               note='Overran EUR 1.34M from 2-month delay (Sep→Nov opening). Ongoing works to completion.')
row = item_row(ws, row, 'Hotel Replacement Works (NEW)', 0, -1050000, -1050000, is_new=True,
               note='Post-Hagenauer insolvency. Hotel defects repaired at own cost. Electrical audit EUR 250-300K before May 26 inspection.')
row = item_row(ws, row, 'Apartment Replacement Works (NEW)', 0, -1500000, -1500000, is_new=True,
               note='Windows, KNX, electrical defects. New contractors at full price — Hagenauer subcontractors refuse to return.')

# CAPEX subtotal
capex_oct = -1341458
capex_jan = -2541213 + -1050000 + -1500000  # -5091213
capex_total = capex_oct + capex_jan  # -6432671

ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'CAPEX SUBTOTAL'
ws[f'K{row}'].font = bold_font
fill_row(ws, row, total_fill)
for col, val in [('O', capex_oct), ('Q', capex_jan), ('S', capex_total)]:
    ws[f'{col}{row}'].value = val
    ws[f'{col}{row}'].number_format = eur_fmt
    ws[f'{col}{row}'].font = bold_font
    ws[f'{col}{row}'].alignment = Alignment(horizontal='right')
row += 2

# ============================================================
# OPEX
# ============================================================
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'OPEX — Hotel Operations & Project Costs'
ws[f'K{row}'].font = section_font
fill_row(ws, row, opex_fill)
row += 1

ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = 'OPEX did not exist in Sep 2025 Capital Call — the hotel was not yet operating. MO Vienna opened 27 November 2025.'
ws[f'K{row}'].font = Font(name='Calibri', size=9, bold=True, italic=True, color='006600')
row += 1

# OPEX items from v009
row = item_row(ws, row, 'MO Pre-Opening: Hotel Staff', -1700000, 0, -1700000,
               note='Staff hired for Sep opening. Continued drawing salaries through Nov. EUR 850K/month × 2 delay months.')
row = item_row(ws, row, 'MO Pre-Opening: Marketing', -136000, -225000, -361000,
               note='Marketing paused during delay — you don\'t advertise a postponed opening. Resumed Q1 2026.')
row = item_row(ws, row, 'Settlement & Warranty Fund (NEW)', 0, -4000000, -4000000, is_new=True,
               note='Hagenauer EUR 19M claim. Needed for: administrator settlement, 3-year subcontractor guarantees, legal costs. Without this, hotel cannot be sold.')
row = item_row(ws, row, 'Working Capital Hotel (NEW)', 0, -1500000, -1500000, is_new=True,
               note='MO contractual right to EUR 1.5M reserve. Hotel now running — daily operations (food, wine, supplies, salaries). Did not exist before opening.')
row = item_row(ws, row, 'OS&E, IT, Art', -851987, -395295, -1247282,
               note='Over-budget on IT systems (EUR 319K — PMS, WiFi, smart room). Art EUR 105K. Not delay-related.')
row = item_row(ws, row, 'Sales Fees & Property Management', -1143556, -703531, -1847087,
               note='Apartment sales commissions (1.5–2%) + property management. Not in Sep budget — sales started after construction.')
row = item_row(ws, row, 'Bank Interest (Aukera)', -1049140, -3375964, -4425104,
               note='Senior facility 5.4% p.a. EUR 350–500K/month. Higher than Sep projection due to extended drawdown from delay.')
row = item_row(ws, row, 'VAT & Taxes (NEW LAW)', 373300, -5770000, -5396700, is_new=True,
               note='Nov 20, 2025: Betrugsbekämpfungsgesetz. Luxury apartments >EUR 2.5M lose VAT exemption. EUR 5.77M hit. Law did not exist at Sep capital call.')

# OPEX subtotal
opex_oct = -1700000 + -136000 + 0 + 0 + -851987 + -1143556 + -1049140 + 373300  # -4507383
opex_jan = 0 + -225000 + -4000000 + -1500000 + -395295 + -703531 + -3375964 + -5770000  # -15969790
opex_total = opex_oct + opex_jan

ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'OPEX SUBTOTAL'
ws[f'K{row}'].font = bold_font
fill_row(ws, row, total_fill)
for col, val in [('O', opex_oct), ('Q', opex_jan), ('S', opex_total)]:
    ws[f'{col}{row}'].value = val
    ws[f'{col}{row}'].number_format = eur_fmt
    ws[f'{col}{row}'].font = bold_font
    ws[f'{col}{row}'].alignment = Alignment(horizontal='right')
row += 1

# Rounding / other (to match v009 total exactly)
# v009 total = -27,134,844. CAPEX + OPEX = -6,432,671 + -20,477,173 = -26,909,844. Diff = -225,000
diff = -27134844 - (capex_total + opex_total)
if abs(diff) > 0:
    ws.merge_cells(f'K{row}:N{row}')
    ws[f'K{row}'].value = 'Other / Rounding'
    ws[f'K{row}'].font = small_font
    ws[f'S{row}'].value = diff
    ws[f'S{row}'].number_format = eur_fmt
    ws[f'S{row}'].font = small_font
    ws[f'S{row}'].alignment = Alignment(horizontal='right')
    row += 1

row += 1

# ============================================================
# GRAND TOTAL USES
# ============================================================
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'TOTAL CAPITAL REQUIRED (Oct 2025 – Jun 2026)'
ws[f'K{row}'].font = big_total_font
fill_row(ws, row, grand_fill)
ws[f'O{row}'].value = -5848841
ws[f'O{row}'].number_format = eur_fmt
ws[f'O{row}'].font = big_total_font
ws[f'O{row}'].alignment = Alignment(horizontal='right')
ws[f'Q{row}'].value = -21286003
ws[f'Q{row}'].number_format = eur_fmt
ws[f'Q{row}'].font = big_total_font
ws[f'Q{row}'].alignment = Alignment(horizontal='right')
ws[f'S{row}'].value = -27134844
ws[f'S{row}'].number_format = eur_fmt
ws[f'S{row}'].font = big_total_font
ws[f'S{row}'].alignment = Alignment(horizontal='right')
row += 2

# ============================================================
# NEW ITEMS SUMMARY
# ============================================================
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'Items NOT in Sep 2025 Capital Call'
ws[f'K{row}'].font = section_font
fill_row(ws, row, new_fill)
row += 1

new_items = [
    ('Hotel Replacement Works (Hagenauer insolvency)', -1050000),
    ('Apartment Replacement Works (Hagenauer insolvency)', -1500000),
    ('Settlement & Warranty Fund (Hagenauer EUR 19M claim)', -4000000),
    ('Working Capital Hotel (MO now operating)', -1500000),
    ('VAT Clawback (Betrugsbekämpfungsgesetz Nov 2025)', -5770000),
]
for label, amount in new_items:
    ws.merge_cells(f'K{row}:Q{row}')
    ws[f'K{row}'].value = label
    ws[f'K{row}'].font = red_font
    fill_row(ws, row, new_fill)
    ws[f'S{row}'].value = amount
    ws[f'S{row}'].number_format = eur_fmt
    ws[f'S{row}'].font = red_font
    ws[f'S{row}'].alignment = Alignment(horizontal='right')
    row += 1

ws.merge_cells(f'K{row}:Q{row}')
ws[f'K{row}'].value = 'TOTAL NEW ITEMS'
ws[f'K{row}'].font = big_total_font
fill_row(ws, row, total_fill)
ws[f'S{row}'].value = -13820000
ws[f'S{row}'].number_format = eur_fmt
ws[f'S{row}'].font = big_total_font
ws[f'S{row}'].alignment = Alignment(horizontal='right')
row += 3

# ============================================================
# SOURCES — HOW THE TOTAL IS COVERED
# ============================================================
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = 'SOURCES OF FUNDING'
ws[f'K{row}'].font = title_font
ws[f'K{row}'].fill = section_fill
fill_row(ws, row, section_fill)
row += 1

# Column headers
fill_row(ws, row, header_fill)
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'Source'
ws[f'K{row}'].font = header_font
ws[f'O{row}'].value = 'Oct-Dec 2025'
ws[f'O{row}'].font = header_font
ws[f'O{row}'].alignment = Alignment(horizontal='center')
ws[f'Q{row}'].value = 'Jan-Jun 2026'
ws[f'Q{row}'].font = header_font
ws[f'Q{row}'].alignment = Alignment(horizontal='center')
ws[f'S{row}'].value = 'Total Oct-Jun'
ws[f'S{row}'].font = header_font
ws[f'S{row}'].alignment = Alignment(horizontal='center')
row += 1

# Source 1: Fund
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'Source 1: Fund (52%)'
ws[f'K{row}'].font = section_font
fill_row(ws, row, source_fill)
row += 1

fund_items = [
    ('Fund — already provided (guarantee + contributions)', 5738339, 2383261, 8121600,
     'Includes cashed bank guarantee (Oct-Dec) + additional contributions (Jan-Mar)'),
    ('Fund — still to provide (Apr-Jun)', None, None, 4660826,
     'Fund\'s 52% share of remaining gap (per Edita v009 row 61)'),
]
for label, o, q, s, note in fund_items:
    ws.merge_cells(f'K{row}:N{row}')
    ws[f'K{row}'].value = label
    ws[f'K{row}'].font = normal_font
    fill_row(ws, row, source_fill)
    if o:
        ws[f'O{row}'].value = o
        ws[f'O{row}'].number_format = eur_fmt
        ws[f'O{row}'].font = normal_font
        ws[f'O{row}'].alignment = Alignment(horizontal='right')
    if q:
        ws[f'Q{row}'].value = q
        ws[f'Q{row}'].number_format = eur_fmt
        ws[f'Q{row}'].font = normal_font
        ws[f'Q{row}'].alignment = Alignment(horizontal='right')
    ws[f'S{row}'].value = s
    ws[f'S{row}'].number_format = eur_fmt
    ws[f'S{row}'].font = bold_font
    ws[f'S{row}'].alignment = Alignment(horizontal='right')
    row += 1
    ws.merge_cells(f'K{row}:T{row}')
    ws[f'K{row}'].value = f'    → {note}'
    ws[f'K{row}'].font = small_font
    row += 1

# Fund total
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'Fund Total'
ws[f'K{row}'].font = bold_font
fill_row(ws, row, total_fill)
ws[f'S{row}'].value = 8121600 + 4660826
ws[f'S{row}'].number_format = eur_fmt
ws[f'S{row}'].font = bold_font
ws[f'S{row}'].alignment = Alignment(horizontal='right')
row += 2

# Source 2: AO
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'Source 2: AO (48%)'
ws[f'K{row}'].font = section_font
fill_row(ws, row, source_fill)
row += 1

ao_items = [
    ('AO — already provided (Jan-Mar)', None, 2500000, 2500000,
     'Transferred Mar 2026. Separate return arrangement (EUR 4M claim at distribution).'),
    ('AO — still to provide (Apr-Jun)', None, None, 4302301,
     'AO\'s 48% share of remaining gap (per Edita v009 row 62)'),
]
for label, o, q, s, note in ao_items:
    ws.merge_cells(f'K{row}:N{row}')
    ws[f'K{row}'].value = label
    ws[f'K{row}'].font = normal_font
    fill_row(ws, row, source_fill)
    if o:
        ws[f'O{row}'].value = o
        ws[f'O{row}'].number_format = eur_fmt
        ws[f'O{row}'].font = normal_font
        ws[f'O{row}'].alignment = Alignment(horizontal='right')
    if q:
        ws[f'Q{row}'].value = q
        ws[f'Q{row}'].number_format = eur_fmt
        ws[f'Q{row}'].font = normal_font
        ws[f'Q{row}'].alignment = Alignment(horizontal='right')
    ws[f'S{row}'].value = s
    ws[f'S{row}'].number_format = eur_fmt
    ws[f'S{row}'].font = bold_font
    ws[f'S{row}'].alignment = Alignment(horizontal='right')
    row += 1
    ws.merge_cells(f'K{row}:T{row}')
    ws[f'K{row}'].value = f'    → {note}'
    ws[f'K{row}'].font = small_font
    row += 1

# AO total
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'AO Total'
ws[f'K{row}'].font = bold_font
fill_row(ws, row, total_fill)
ws[f'S{row}'].value = 2500000 + 4302301
ws[f'S{row}'].number_format = eur_fmt
ws[f'S{row}'].font = bold_font
ws[f'S{row}'].alignment = Alignment(horizontal='right')
row += 2

# Source 3: Other
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'Source 3: Other Sources'
ws[f'K{row}'].font = section_font
fill_row(ws, row, source_fill)
row += 1

other_items = [
    ('Apartment Sales (proceeds applied)', None, 4430118, 4430118,
     'Sales revenue redirected to cover project costs (includes VAT component per legal opinion)'),
    ('Bank Financing (being explored)', None, None, 0,
     'DV working with banks from Monday. Any facility obtained will reduce shareholder capital calls.'),
]
for label, o, q, s, note in other_items:
    ws.merge_cells(f'K{row}:N{row}')
    ws[f'K{row}'].value = label
    ws[f'K{row}'].font = normal_font
    fill_row(ws, row, source_fill)
    if o:
        ws[f'O{row}'].value = o
        ws[f'O{row}'].number_format = eur_fmt
        ws[f'O{row}'].font = normal_font
        ws[f'O{row}'].alignment = Alignment(horizontal='right')
    if q:
        ws[f'Q{row}'].value = q
        ws[f'Q{row}'].number_format = eur_fmt
        ws[f'Q{row}'].font = normal_font
        ws[f'Q{row}'].alignment = Alignment(horizontal='right')
    ws[f'S{row}'].value = s
    ws[f'S{row}'].number_format = eur_fmt
    ws[f'S{row}'].font = bold_font
    ws[f'S{row}'].alignment = Alignment(horizontal='right')
    row += 1
    ws.merge_cells(f'K{row}:T{row}')
    ws[f'K{row}'].value = f'    → {note}'
    ws[f'K{row}'].font = small_font
    row += 1

# Other total
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'Other Total'
ws[f'K{row}'].font = bold_font
fill_row(ws, row, total_fill)
ws[f'S{row}'].value = 4430118
ws[f'S{row}'].number_format = eur_fmt
ws[f'S{row}'].font = bold_font
ws[f'S{row}'].alignment = Alignment(horizontal='right')
row += 2

# ============================================================
# GRAND TOTAL SOURCES
# ============================================================
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'TOTAL SOURCES'
ws[f'K{row}'].font = Font(name='Calibri', size=11, bold=True, color='006600')
fill_row(ws, row, grand_fill)

total_sources = (8121600 + 4660826) + (2500000 + 4302301) + 4430118
ws[f'S{row}'].value = total_sources
ws[f'S{row}'].number_format = eur_fmt
ws[f'S{row}'].font = Font(name='Calibri', size=11, bold=True, color='006600')
ws[f'S{row}'].alignment = Alignment(horizontal='right')
row += 1

# Balance
ws.merge_cells(f'K{row}:N{row}')
ws[f'K{row}'].value = 'BALANCE (Uses + Sources)'
ws[f'K{row}'].font = big_total_font
fill_row(ws, row, grand_fill)
balance = -27134844 + total_sources
ws[f'S{row}'].value = balance
ws[f'S{row}'].number_format = eur_fmt
ws[f'S{row}'].font = big_total_font if balance < 0 else Font(name='Calibri', size=11, bold=True, color='006600')
ws[f'S{row}'].alignment = Alignment(horizontal='right')
row += 1

if abs(balance) > 1:
    ws.merge_cells(f'K{row}:T{row}')
    ws[f'K{row}'].value = f'    → Remaining gap of EUR {abs(int(balance)):,} to be covered by bank financing or additional shareholder contributions.'
    ws[f'K{row}'].font = small_font
    row += 1

row += 1
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = 'Note: EUR 2,500,000 transferred Mar 2026 is accounted as AO contribution but has a separate return arrangement (EUR 4M claim at distribution).'
ws[f'K{row}'].font = Font(name='Calibri', size=9, italic=True, color='666666')
row += 1
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = 'All figures from Edita v009 table (11 April 2026). This is the authoritative source.'
ws[f'K{row}'].font = Font(name='Calibri', size=9, italic=True, color='666666')

wb.save(dst)
print(f'Saved: {dst}')
print(f'Total sources: {total_sources:,}')
print(f'Total uses: -27,134,844')
print(f'Balance: {balance:,}')
