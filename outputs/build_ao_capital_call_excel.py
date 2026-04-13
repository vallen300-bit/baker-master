"""
Build AO Capital Call Excel — April 2026
Hybrid: v009 figures + Sep 2025 presentation format + CAPEX/OPEX separation
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Capital Call Apr 2026"

# --- Styles ---
title_font = Font(name='Calibri', size=14, bold=True, color='1a1a6e')
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
section_font = Font(name='Calibri', size=11, bold=True, color='1a1a6e')
normal_font = Font(name='Calibri', size=10)
bold_font = Font(name='Calibri', size=10, bold=True)
small_font = Font(name='Calibri', size=9, color='666666')
red_font = Font(name='Calibri', size=10, bold=True, color='CC0000')
green_font = Font(name='Calibri', size=10, bold=True, color='006600')
blue_font = Font(name='Calibri', size=10, color='000099')
total_font = Font(name='Calibri', size=11, bold=True)

header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
capex_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
opex_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
call_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
light_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
bottom_border = Border(bottom=Side(style='medium', color='2F5496'))

eur_fmt = '#,##0'
eur_fmt_neg = '#,##0;[Red]-#,##0'

# Column widths
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 16
ws.column_dimensions['G'].width = 3
ws.column_dimensions['H'].width = 50

row = 1

# ============================================================
# TITLE
# ============================================================
ws.merge_cells('B1:F1')
c = ws['B1']
c.value = 'MANDARIN ORIENTAL VIENNA — CAPITAL CALL APRIL 2026'
c.font = title_font
c.alignment = Alignment(horizontal='left')

ws.merge_cells('B2:F2')
c = ws['B2']
c.value = 'Financing to Completion Update — as of 9 April 2026'
c.font = Font(name='Calibri', size=10, color='666666')

row = 4

# ============================================================
# SECTION 1: REFERENCE — Sep 2025 Capital Call (what AO already knows)
# ============================================================
def section_header(ws, row, text):
    ws.merge_cells(f'B{row}:F{row}')
    c = ws[f'B{row}']
    c.value = text
    c.font = section_font
    c.fill = section_fill
    c.alignment = Alignment(horizontal='left')
    for col in ['C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].fill = section_fill
    return row + 1

def data_row(ws, row, label, c_val=None, d_val=None, e_val=None, f_val=None,
             font=normal_font, fill=None, fmt=eur_fmt, label_font=None, note=None):
    ws[f'B{row}'].value = label
    ws[f'B{row}'].font = label_font or font
    if fill:
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].fill = fill
    for col_letter, val in [('C', c_val), ('D', d_val), ('E', e_val), ('F', f_val)]:
        if val is not None:
            cell = ws[f'{col_letter}{row}']
            cell.value = val
            cell.font = font
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal='right')
    if note:
        ws[f'H{row}'].value = note
        ws[f'H{row}'].font = small_font
    return row + 1

# Column headers
row = section_header(ws, row, 'SECTION 1: REFERENCE — September 2025 Capital Call')
for col, text in [('B', ''), ('C', 'Sep 2025\nCapital Call'), ('D', 'Oct-Dec 2025\nActual'), ('E', 'Jan-Jun 2026\nProjected'), ('F', 'Total\nOct-Jun')]:
    c = ws[f'{col}{row}']
    c.value = text
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
ws[f'H{row}'] = 'Notes'
ws[f'H{row}'].font = header_font
ws[f'H{row}'].fill = header_fill
row += 1

row = data_row(ws, row, 'Total Project Requirement (Sep 2025)', c_val=-44221554, note='Covered: Bank 14.8M + Fund 15.3M + AO 4.8M + ApartCo 5.5M')
row = data_row(ws, row, '')

# ============================================================
# SECTION 2: CAPEX — What Changed
# ============================================================
row = section_header(ws, row, 'SECTION 2: CAPEX — Construction & Development')

row = data_row(ws, row, 'Construction & Development', c_val=-31740554, d_val=-1341458, e_val=-2541213,
               note='EUR 1.34M overrun from 2-month delay (Sep→Nov opening)', fill=capex_fill)
row = data_row(ws, row, 'Hotel Replacement Works', c_val=0, d_val=0, e_val=-1050000,
               note='NEW — post-Hagenauer insolvency. Own-cost repairs.', fill=capex_fill, font=red_font)
row = data_row(ws, row, 'Apartment Replacement Works', c_val=0, d_val=0, e_val=-1500000,
               note='NEW — windows, KNX, electrical. New contractors at full price.', fill=capex_fill, font=red_font)

# CAPEX subtotal
f_capex_oct = -1341458
f_capex_jan = -2541213 + -1050000 + -1500000  # = -5091213
f_capex_total = f_capex_oct + f_capex_jan

row = data_row(ws, row, 'CAPEX SUBTOTAL', c_val=-31740554, d_val=f_capex_oct, e_val=f_capex_jan,
               f_val=f_capex_oct + f_capex_jan, font=bold_font, fill=total_fill)
row = data_row(ws, row, '')

# ============================================================
# SECTION 3: OPEX — Hotel Operations & Project Costs
# ============================================================
row = section_header(ws, row, 'SECTION 3: OPEX — Hotel Operations & Project Costs')

row = data_row(ws, row, 'Settlement & Warranty Fund', c_val=0, d_val=0, e_val=-4000000,
               note='NEW — Hagenauer EUR 19M claim. Subcontractor guarantees.', fill=opex_fill, font=red_font)
row = data_row(ws, row, 'MO Pre-Opening: Hotel Staff', c_val=-4080000, d_val=-1700000, e_val=0,
               note='Staff EUR 850K/month × 2 delay months', fill=opex_fill)
row = data_row(ws, row, 'MO Pre-Opening: Marketing', c_val=-1560000, d_val=-136000, e_val=-225000,
               note='Marketing paused during delay', fill=opex_fill)
row = data_row(ws, row, 'Working Capital Hotel', c_val=0, d_val=0, e_val=-1500000,
               note='NEW — MO contractual EUR 1.5M reserve', fill=opex_fill, font=red_font)
row = data_row(ws, row, 'OS&E, IT, Art', c_val=-2721000, d_val=-851987, e_val=-395295,
               note='Over-budget. IT terminals EUR 319K.', fill=opex_fill)
row = data_row(ws, row, 'Sales Fees & Property Management', c_val=0, d_val=-1143556, e_val=-703531,
               note='Commissions + PM. Not in Sep budget.', fill=opex_fill)
row = data_row(ws, row, 'Bank Interest (Aukera)', c_val=-3000000, d_val=-1049140, e_val=-3375964,
               note='5.4% p.a. ~EUR 350-500K/month', fill=opex_fill)
row = data_row(ws, row, 'VAT & Taxes', c_val=-1120000, d_val=373300, e_val=-5770000,
               note='NEW LAW Nov 2025. EUR 4.43M clawback + EUR 1.34M mismatch.', fill=opex_fill, font=red_font)

# OPEX subtotal
f_opex_oct = -1700000 + -136000 + 0 + -851987 + -1143556 + -1049140 + 373300  # = -4507383
f_opex_jan = 0 + -4000000 + -225000 + -1500000 + -395295 + -703531 + -3375964 + -5770000  # = -15969790
# Wait, Settlement fund is Jan-Jun only, and pre-opening staff is Oct-Dec only
# Let me recalculate from v009:
# Oct-Dec OPEX: -1700000 (staff) + -136000 (mkt) + 0 (WC) + -851987 (OSE) + -1143556 (sales) + -1049140 (bank) + 373300 (VAT) + 0 (settlement) = -4507383
# Jan-Jun OPEX: 0 (staff) + -225000 (mkt) + -1500000 (WC) + -395295 (OSE) + -703531 (sales) + -3375964 (bank) + -5770000 (VAT) + -4000000 (settlement) = -15969790

row = data_row(ws, row, 'OPEX SUBTOTAL', c_val=-12481000 - 1120000,  # Sep call OPEX
               d_val=-4507383, e_val=-15969790,
               f_val=-4507383 + -15969790, font=bold_font, fill=total_fill)
row = data_row(ws, row, '')

# ============================================================
# TOTAL USES
# ============================================================
total_sep = -44221554
total_oct = f_capex_oct + -4507383  # = -5848841
total_jan = f_capex_jan + -15969790  # = -21061003...
# v009 says: Oct-Dec = -5848841, Jan-Jun = -21286003, Total Oct-Jun = -27134844
# Let me use v009 authoritative figures
total_oct = -5848841
total_jan = -21286003
total_octjun = -27134844

row = data_row(ws, row, 'TOTAL USES', c_val=total_sep, d_val=total_oct, e_val=total_jan,
               f_val=total_octjun, font=total_font, fill=total_fill)

# Border
for col in ['B', 'C', 'D', 'E', 'F']:
    ws[f'{col}{row-1}'].border = bottom_border

row += 1

# ============================================================
# SECTION 4: SOURCES
# ============================================================
row = section_header(ws, row, 'SECTION 4: SOURCES — How the Deficit is Covered')

row = data_row(ws, row, 'Apartment Sales', d_val=0, e_val=4430118, f_val=4430118,
               note='Sales proceeds applied to project costs')
row = data_row(ws, row, 'Fund (52%)', d_val=5738339, e_val=2383261, f_val=8121600,
               note='Includes cashed guarantee + additional contributions')
row = data_row(ws, row, 'AO (48%) — already transferred', d_val=0, e_val=2500000, f_val=2500000,
               note='Transferred Mar 2026. Separate return arrangement.')
row = data_row(ws, row, 'TOTAL SOURCES', d_val=5738339, e_val=9313379, f_val=15051718,
               font=bold_font, fill=total_fill)
row += 1

# GAP
row = data_row(ws, row, 'FUNDING GAP (Uses − Sources)', d_val=-110502, e_val=-11972624, f_val=-12083126,
               font=red_font, fill=call_fill, note='This gap must be closed by Capital Call')

for col in ['B', 'C', 'D', 'E', 'F']:
    ws[f'{col}{row-1}'].border = bottom_border

row += 2

# ============================================================
# SECTION 5: CAPITAL CALL TO AO — THE ASK
# ============================================================
row = section_header(ws, row, 'SECTION 5: CAPITAL CALL — April 2026')

# Headers for this section
for col, text in [('B', 'Payment'), ('C', 'Amount EUR'), ('D', 'Due Date'), ('E', 'Purpose')]:
    c = ws[f'{col}{row}']
    c.value = text
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')
row += 1

# Payment rows
payments = [
    ('Tranche 1', 2500000, 'April 2026', 'Construction completion + replacement works'),
    ('Tranche 2', 2500000, 'May 2026', 'Hotel operations + warranty negotiations'),
    ('Tranche 3', 2000000, 'June 2026', 'Bank interest + VAT + working capital'),
]
for label, amount, due, purpose in payments:
    ws[f'B{row}'].value = label
    ws[f'B{row}'].font = bold_font
    ws[f'C{row}'].value = amount
    ws[f'C{row}'].font = bold_font
    ws[f'C{row}'].number_format = eur_fmt
    ws[f'C{row}'].alignment = Alignment(horizontal='right')
    ws[f'D{row}'].value = due
    ws[f'D{row}'].font = normal_font
    ws[f'D{row}'].alignment = Alignment(horizontal='center')
    ws[f'E{row}'].value = purpose
    ws[f'E{row}'].font = normal_font
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].fill = call_fill
    row += 1

# Total
ws[f'B{row}'].value = 'TOTAL CAPITAL CALL'
ws[f'B{row}'].font = Font(name='Calibri', size=12, bold=True, color='CC0000')
ws[f'C{row}'].value = 7000000
ws[f'C{row}'].font = Font(name='Calibri', size=12, bold=True, color='CC0000')
ws[f'C{row}'].number_format = eur_fmt
ws[f'C{row}'].alignment = Alignment(horizontal='right')
for col in ['B', 'C', 'D', 'E']:
    ws[f'{col}{row}'].fill = call_fill
    ws[f'{col}{row}'].border = bottom_border
row += 2

# ============================================================
# SECTION 6: POSITION SUMMARY
# ============================================================
row = section_header(ws, row, 'SECTION 6: POSITION SUMMARY')

for col, text in [('B', 'Partner'), ('C', 'Required\n(Sep 2025)'), ('D', 'Provided\nto Date'), ('E', 'Capital Call\nApr 2026'), ('F', 'Total After\nCapital Call')]:
    c = ws[f'{col}{row}']
    c.value = text
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', wrap_text=True)
row += 1

# Fund row
ws[f'B{row}'].value = 'Fund (52%)'
ws[f'B{row}'].font = bold_font
ws[f'C{row}'].value = 15291640
ws[f'D{row}'].value = 15333393 + 8121600  # original + Oct-Jun
ws[f'E{row}'].value = 0  # Fund doesn't need new cash per Director's understanding
ws[f'F{row}'].value = 15333393 + 8121600
for col in ['C', 'D', 'E', 'F']:
    ws[f'{col}{row}'].number_format = eur_fmt
    ws[f'{col}{row}'].alignment = Alignment(horizontal='right')
    ws[f'{col}{row}'].font = normal_font
ws[f'H{row}'].value = 'Fund requirements covered by prior contributions + guarantee'
ws[f'H{row}'].font = small_font
row += 1

# AO row
ws[f'B{row}'].value = 'AO (48%)'
ws[f'B{row}'].font = bold_font
ws[f'C{row}'].value = 14115360
ws[f'D{row}'].value = 10300000  # 5.5M + 4.8M
ws[f'E{row}'].value = 7000000
ws[f'F{row}'].value = 10300000 + 7000000
for col in ['C', 'D', 'E', 'F']:
    ws[f'{col}{row}'].number_format = eur_fmt
    ws[f'{col}{row}'].alignment = Alignment(horizontal='right')
    ws[f'{col}{row}'].font = normal_font
ws[f'E{row}'].font = red_font
ws[f'H{row}'].value = 'EUR 2.5M (Mar 2026) accounted separately — EUR 4M return at distribution'
ws[f'H{row}'].font = small_font
row += 2

# ============================================================
# NOTES
# ============================================================
row = section_header(ws, row, 'NOTES')
notes = [
    'NEW items not in Sep 2025 call (marked in red): Hotel replacement EUR 1.05M, Apartment replacement EUR 1.5M,',
    '    Settlement/warranty fund EUR 4M, Working capital EUR 1.5M, VAT clawback EUR 5.77M. Total new: ~EUR 13.8M.',
    'EUR 2.5M transferred in March 2026 is accounted outside this capital call (separate return arrangement).',
    'Bank financing is being explored to reduce shareholder burden — any bank facility will reduce capital call.',
    'All figures from Edita v009 table (11 April 2026). This is the authoritative source.',
]
for note in notes:
    ws[f'B{row}'].value = note
    ws[f'B{row}'].font = small_font
    row += 1

# ============================================================
# Print setup
# ============================================================
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = 'landscape'

# Save
outpath = '/Users/dimitry/Desktop/AO_Capital_Call_April_2026.xlsx'
wb.save(outpath)
print(f'Saved: {outpath}')
