"""
Build AO explanations in v009 — Baker version 3.
Sources section mirrors Edita rows 58-63 structure exactly.
"""
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

src = "/Users/dimitry/Desktop/009_MOVIE_AO_Baker_Version_20260411 Baker version 3.xlsx"
dst = "/Users/dimitry/Desktop/009_MOVIE_AO_Baker_Version_20260411 Baker version 3.xlsx"

wb = openpyxl.load_workbook(dst)
ws = wb['AO ']

# --- Styles ---
title_font = Font(name='Calibri', size=12, bold=True, color='1a1a6e')
header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
section_font = Font(name='Calibri', size=10, bold=True, color='1a1a6e')
normal_font = Font(name='Calibri', size=10)
bold_font = Font(name='Calibri', size=10, bold=True)
small_font = Font(name='Calibri', size=9, color='555555')
red_font = Font(name='Calibri', size=10, bold=True, color='CC0000')
big_total_font = Font(name='Calibri', size=11, bold=True, color='CC0000')
green_bold = Font(name='Calibri', size=11, bold=True, color='006600')

header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
source_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
grand_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
new_fill = PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid')

COLS = ['K','L','M','N','O','P','Q','R','S','T']
eur_fmt = '#,##0'

def fill_row(ws, row, fill, cols=COLS):
    for c in cols:
        ws[f'{c}{row}'].fill = fill

def num_cell(ws, col, row, val, font=normal_font):
    ws[f'{col}{row}'].value = val
    ws[f'{col}{row}'].number_format = eur_fmt
    ws[f'{col}{row}'].font = font
    ws[f'{col}{row}'].alignment = Alignment(horizontal='right')

# --- Find where Sources section starts (after the USES total and new items summary) ---
# Scan for "SOURCES OF FUNDING" or "Source 1" to find the old section
start_clear = None
for r in range(71, 150):
    val = ws.cell(row=r, column=11).value  # column K
    if val and 'SOURCES OF FUNDING' in str(val):
        start_clear = r
        break

if start_clear is None:
    # Find "Items NOT in Sep" section end, then the Sources start after
    for r in range(71, 150):
        val = ws.cell(row=r, column=11).value
        if val and 'TOTAL NEW ITEMS' in str(val):
            start_clear = r + 2  # 2 rows after
            break

if start_clear is None:
    start_clear = 105  # fallback

print(f"Clearing from row {start_clear} to 150")

# Unmerge cells in the range
merges_to_remove = []
for merge in ws.merged_cells.ranges:
    if merge.min_row >= start_clear and merge.min_row <= 150:
        merges_to_remove.append(merge)
for merge in merges_to_remove:
    ws.unmerge_cells(str(merge))

# Clear
for r in range(start_clear, 151):
    for c in range(11, 21):
        cell = ws.cell(row=r, column=c)
        cell.value = None
        cell.font = Font(name='Calibri', size=10)
        cell.fill = PatternFill(fill_type=None)
        cell.border = Border()
        cell.alignment = Alignment()
        cell.number_format = 'General'

row = start_clear

# ============================================================
# SOURCES — mirroring Edita rows 58-63 structure
# ============================================================
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = 'SOURCES OF FUNDING — How the EUR 27,134,844 is Covered'
ws[f'K{row}'].font = title_font
ws[f'K{row}'].fill = section_fill
fill_row(ws, row, section_fill)
row += 2

# Period headers — matching Edita's column structure
fill_row(ws, row, header_fill)
ws[f'K{row}'].value = 'Source'
ws[f'K{row}'].font = header_font
ws[f'L{row}'].value = 'by 30.06.2025'
ws[f'L{row}'].font = header_font
ws[f'L{row}'].alignment = Alignment(horizontal='center')
ws.merge_cells(f'M{row}:N{row}')
ws[f'M{row}'].value = 'by 30.09.2025'
ws[f'M{row}'].font = header_font
ws[f'M{row}'].alignment = Alignment(horizontal='center')
ws[f'N{row}'].fill = header_fill
ws[f'O{row}'].value = 'Oct-Dec 2025'
ws[f'O{row}'].font = header_font
ws[f'O{row}'].alignment = Alignment(horizontal='center')
ws[f'P{row}'].value = 'Provided'
ws[f'P{row}'].font = header_font
ws[f'P{row}'].alignment = Alignment(horizontal='center')
ws[f'Q{row}'].value = 'Jan-Mar 2026'
ws[f'Q{row}'].font = header_font
ws[f'Q{row}'].alignment = Alignment(horizontal='center')
ws[f'R{row}'].value = 'Provided'
ws[f'R{row}'].font = header_font
ws[f'R{row}'].alignment = Alignment(horizontal='center')
ws[f'S{row}'].value = 'Apr-Jun 2026'
ws[f'S{row}'].font = header_font
ws[f'S{row}'].alignment = Alignment(horizontal='center')
ws[f'T{row}'].value = 'To Provide'
ws[f'T{row}'].font = header_font
ws[f'T{row}'].alignment = Alignment(horizontal='center')
row += 1

# Sub-headers
fill_row(ws, row, PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'))
ws[f'L{row}'].value = 'Required'
ws[f'L{row}'].font = Font(name='Calibri', size=8, color='333333')
ws[f'L{row}'].alignment = Alignment(horizontal='center')
ws.merge_cells(f'M{row}:N{row}')
ws[f'M{row}'].value = 'Loans Provided'
ws[f'M{row}'].font = Font(name='Calibri', size=8, color='333333')
ws[f'M{row}'].alignment = Alignment(horizontal='center')
ws[f'O{row}'].value = 'Required Addl'
ws[f'O{row}'].font = Font(name='Calibri', size=8, color='333333')
ws[f'O{row}'].alignment = Alignment(horizontal='center')
ws[f'P{row}'].value = 'Loans Provided'
ws[f'P{row}'].font = Font(name='Calibri', size=8, color='333333')
ws[f'P{row}'].alignment = Alignment(horizontal='center')
ws[f'Q{row}'].value = 'Required Addl'
ws[f'Q{row}'].font = Font(name='Calibri', size=8, color='333333')
ws[f'Q{row}'].alignment = Alignment(horizontal='center')
ws[f'R{row}'].value = 'Loans Provided'
ws[f'R{row}'].font = Font(name='Calibri', size=8, color='333333')
ws[f'R{row}'].alignment = Alignment(horizontal='center')
ws[f'S{row}'].value = 'Required Addl'
ws[f'S{row}'].font = Font(name='Calibri', size=8, color='333333')
ws[f'S{row}'].alignment = Alignment(horizontal='center')
ws[f'T{row}'].value = 'To be Provided'
ws[f'T{row}'].font = Font(name='Calibri', size=8, color='333333')
ws[f'T{row}'].alignment = Alignment(horizontal='center')
row += 1

# --- SHAREHOLDER ROWS (from Edita rows 61-63) ---

# Fund (52%)
ws[f'K{row}'].value = 'Fund (52%)'
ws[f'K{row}'].font = bold_font
fill_row(ws, row, source_fill)
num_cell(ws, 'L', row, 15291640, bold_font)     # Required by Jun 2025
num_cell(ws, 'M', row, 15333393, bold_font)      # Provided by Sep 2025
num_cell(ws, 'O', row, 3041397, normal_font)      # Required Oct-Dec
num_cell(ws, 'P', row, 5738339, bold_font)        # Provided Oct-Dec (guarantee cashed)
num_cell(ws, 'Q', row, 2539296, normal_font)      # Required Jan-Mar
num_cell(ws, 'R', row, 2383261, bold_font)        # Provided Jan-Mar
num_cell(ws, 'S', row, None)                       # Required Apr-Jun (part of total)
num_cell(ws, 'T', row, 4660826, red_font)         # To be provided Apr-Jun
row += 1

# Explanation
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = '    → Oct-Dec: Fund provided EUR 5.74M (cashed bank guarantee) against EUR 3.04M required — covered AO shortfall. Jan-Mar: EUR 2.38M provided. Apr-Jun: EUR 4.66M still to provide.'
ws[f'K{row}'].font = small_font
row += 1

# AO (48%)
ws[f'K{row}'].value = 'Shareholder AO (48%)'
ws[f'K{row}'].font = bold_font
fill_row(ws, row, source_fill)
num_cell(ws, 'L', row, 14115360, bold_font)      # Required by Jun 2025
num_cell(ws, 'M', row, 4800000, bold_font)        # Provided by Sep 2025
num_cell(ws, 'O', row, 2807444, normal_font)      # Required Oct-Dec
num_cell(ws, 'P', row, 0, red_font)               # Provided Oct-Dec = 0
num_cell(ws, 'Q', row, 2343965, normal_font)      # Required Jan-Mar
num_cell(ws, 'R', row, 2500000, bold_font)        # Provided Jan-Mar
num_cell(ws, 'S', row, None)
num_cell(ws, 'T', row, 4302301, red_font)         # To be provided Apr-Jun
row += 1

# Explanation
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = '    → Oct-Dec: EUR 2.81M required, EUR 0 provided. Jan-Mar: EUR 2.5M provided (separate return arrangement — EUR 4M claim at distribution). Apr-Jun: EUR 4.30M still to provide.'
ws[f'K{row}'].font = small_font
row += 1

# Total shareholders
ws[f'K{row}'].value = 'Shareholders Total'
ws[f'K{row}'].font = bold_font
fill_row(ws, row, total_fill)
num_cell(ws, 'L', row, 29407000, bold_font)
num_cell(ws, 'M', row, 20133393, bold_font)
num_cell(ws, 'O', row, 5848841, bold_font)
num_cell(ws, 'P', row, 5738339, bold_font)  # Fund only in Oct-Dec
num_cell(ws, 'Q', row, 4883261, bold_font)
num_cell(ws, 'R', row, 4883261, bold_font)
num_cell(ws, 'S', row, 8963126, bold_font)
num_cell(ws, 'T', row, 8963126, bold_font)
row += 2

# --- OTHER SOURCES ---
ws[f'K{row}'].value = 'Other Sources'
ws[f'K{row}'].font = section_font
fill_row(ws, row, section_fill)
row += 1

# Bank finance (original)
ws[f'K{row}'].value = 'Bank Finance (Aukera — original)'
ws[f'K{row}'].font = normal_font
fill_row(ws, row, source_fill)
num_cell(ws, 'L', row, 14814554, bold_font)  # by Jun 2025
ws.merge_cells(f'M{row}:N{row}')
ws[f'M{row}'].value = 'Fully drawn'
ws[f'M{row}'].font = small_font
ws[f'M{row}'].alignment = Alignment(horizontal='center')
row += 1

# Apartment sales
ws[f'K{row}'].value = 'Apartment Sales Proceeds'
ws[f'K{row}'].font = normal_font
fill_row(ws, row, source_fill)
num_cell(ws, 'R', row, 4430118, bold_font)  # Jan-Mar received
num_cell(ws, 'T', row, 3120000, bold_font)  # Apr-Jun expected
row += 1
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = '    → Sales proceeds applied to project costs. EUR 4.43M received. EUR 3.12M expected Apr-Jun (from Edita v009 capital call column).'
ws[f'K{row}'].font = small_font
row += 1

# Bank financing (new — being explored)
ws[f'K{row}'].value = 'Additional Bank Financing (being explored)'
ws[f'K{row}'].font = normal_font
fill_row(ws, row, source_fill)
num_cell(ws, 'T', row, 0, normal_font)
row += 1
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = '    → Working with banks to secure additional facility. Any amount obtained reduces shareholder capital calls.'
ws[f'K{row}'].font = small_font
row += 2

# ============================================================
# SUMMARY: USES vs SOURCES
# ============================================================
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = 'SUMMARY'
ws[f'K{row}'].font = title_font
ws[f'K{row}'].fill = section_fill
fill_row(ws, row, section_fill)
row += 1

summary_items = [
    ('Total Capital Required (Oct 2025 – Jun 2026)', -27134844, big_total_font),
    ('Already Provided: Fund (Oct-Dec + Jan-Mar)', 5738339 + 2383261, bold_font),
    ('Already Provided: AO (Jan-Mar)', 2500000, bold_font),
    ('Already Provided: Apartment Sales', 4430118, bold_font),
    ('Total Already Provided', 5738339 + 2383261 + 2500000 + 4430118, green_bold),
    ('', None, normal_font),
    ('Remaining Gap', -27134844 + 5738339 + 2383261 + 2500000 + 4430118, big_total_font),
    ('', None, normal_font),
    ('To be covered by:', None, section_font),
    ('    Fund (52%) — Apr-Jun', 4660826, bold_font),
    ('    AO (48%) — Apr-Jun', 4302301, bold_font),
    ('    Apartment Sales (expected Apr-Jun)', 3120000, bold_font),
    ('    Total to be provided', 4660826 + 4302301 + 3120000, green_bold),
]

for label, val, font in summary_items:
    ws[f'K{row}'].value = label
    ws[f'K{row}'].font = font
    if val is not None:
        num_cell(ws, 'S', row, val, font)
    row += 1

row += 1

# Verification
total_provided = 5738339 + 2383261 + 2500000 + 4430118
total_to_provide = 4660826 + 4302301 + 3120000
balance = -27134844 + total_provided + total_to_provide

ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = f'Verification: -27,134,844 + {total_provided:,} (provided) + {total_to_provide:,} (to provide) = {balance:,}'
ws[f'K{row}'].font = Font(name='Calibri', size=9, italic=True, color='006600' if abs(balance) < 100 else 'CC0000')
row += 2

# Notes
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = 'Note: EUR 2.5M transferred Mar 2026 (AO) has separate return arrangement — EUR 4M claim at distribution. Accounted as AO contribution above.'
ws[f'K{row}'].font = Font(name='Calibri', size=9, italic=True, color='666666')
row += 1
ws.merge_cells(f'K{row}:T{row}')
ws[f'K{row}'].value = 'All figures from Edita v009 table (11 April 2026). Rows 58-63 are the authoritative shareholder allocation.'
ws[f'K{row}'].font = Font(name='Calibri', size=9, italic=True, color='666666')

wb.save(dst)
print(f'Saved: {dst}')
print(f'Total provided: {total_provided:,}')
print(f'Total to provide: {total_to_provide:,}')
print(f'Balance: {balance:,}')
