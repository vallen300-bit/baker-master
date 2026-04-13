"""
Fix Sources section — add Disbalance (U) and Adjusted To Provide (V) columns.
AO to provide = 6,900,669 (not 4,302,301).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

dst = "/Users/dimitry/Desktop/009_MOVIE_AO_Baker_Version_20260411 Baker version 3.xlsx"
wb = openpyxl.load_workbook(dst)
ws = wb['AO ']

# Styles
bold_font = Font(name='Calibri', size=10, bold=True)
normal_font = Font(name='Calibri', size=10)
small_font = Font(name='Calibri', size=9, color='555555')
red_font = Font(name='Calibri', size=10, bold=True, color='CC0000')
green_bold = Font(name='Calibri', size=11, bold=True, color='006600')
big_total_font = Font(name='Calibri', size=11, bold=True, color='CC0000')
header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
section_font = Font(name='Calibri', size=10, bold=True, color='1a1a6e')

header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
source_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
disb_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')  # orange tint for disbalance

eur_fmt = '#,##0'

def num_cell(ws, col, row, val, font=normal_font):
    ws[f'{col}{row}'].value = val
    ws[f'{col}{row}'].number_format = eur_fmt
    ws[f'{col}{row}'].font = font
    ws[f'{col}{row}'].alignment = Alignment(horizontal='right')

# Find the Sources header row and shareholder rows
# Scan for "Fund (52%)" in our additions (row > 70)
fund_row = None
ao_row = None
total_sh_row = None
header_row = None

for r in range(100, 150):
    val = ws.cell(row=r, column=11).value  # col K
    if val and 'Fund (52%)' in str(val) and r > 100:
        fund_row = r
    elif val and 'Shareholder AO (48%)' in str(val) and r > 100:
        ao_row = r
    elif val and 'Shareholders Total' in str(val) and r > 100:
        total_sh_row = r
    elif val and 'Source' in str(val) and r > 100:
        header_row = r

print(f"Found: header={header_row}, fund={fund_row}, ao={ao_row}, total={total_sh_row}")

if not all([fund_row, ao_row, total_sh_row]):
    # Fallback: search more broadly
    for r in range(70, 150):
        val = ws.cell(row=r, column=11).value
        if val:
            if 'Fund (52%)' in str(val) and fund_row is None:
                fund_row = r
            elif '48%' in str(val) and 'AO' in str(val) and ao_row is None:
                ao_row = r
            elif 'Shareholders Total' in str(val) and total_sh_row is None:
                total_sh_row = r
    print(f"Fallback found: fund={fund_row}, ao={ao_row}, total={total_sh_row}")

# Find the header row (1 or 2 rows before fund_row)
if fund_row:
    for candidate in [fund_row - 3, fund_row - 2, fund_row - 1]:
        val = ws.cell(row=candidate, column=20).value  # col T
        if val and 'Provide' in str(val):
            header_row = candidate
            break

print(f"Header row: {header_row}")

# ADD columns U and V headers
if header_row:
    # U header
    ws[f'U{header_row}'].value = 'Disbalance'
    ws[f'U{header_row}'].font = header_font
    ws[f'U{header_row}'].fill = PatternFill(start_color='E65100', end_color='E65100', fill_type='solid')
    ws[f'U{header_row}'].alignment = Alignment(horizontal='center')

    # V header
    ws[f'V{header_row}'].value = 'To Provide\n(Adjusted)'
    ws[f'V{header_row}'].font = header_font
    ws[f'V{header_row}'].fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
    ws[f'V{header_row}'].alignment = Alignment(horizontal='center', wrap_text=True)

    # Sub-header row (header_row + 1)
    sub_row = header_row + 1
    ws[f'U{sub_row}'].value = 'Over/Under'
    ws[f'U{sub_row}'].font = Font(name='Calibri', size=8, color='333333')
    ws[f'U{sub_row}'].alignment = Alignment(horizontal='center')
    ws[f'V{sub_row}'].value = 'After Disbalance'
    ws[f'V{sub_row}'].font = Font(name='Calibri', size=8, color='333333')
    ws[f'V{sub_row}'].alignment = Alignment(horizontal='center')

# Fund row — add U and V
if fund_row:
    ws[f'U{fund_row}'].fill = disb_fill
    num_cell(ws, 'U', fund_row, 2598368, normal_font)  # Fund over-provided
    ws[f'U{fund_row}'].fill = disb_fill

    ws[f'V{fund_row}'].fill = source_fill
    num_cell(ws, 'V', fund_row, 2062458, bold_font)  # Fund adjusted: pays LESS

    # Update T to keep but now V is the real number
    # T stays as 4,660,826 (pure 52%)

# AO row — add U and V
if ao_row:
    ws[f'U{ao_row}'].fill = disb_fill
    num_cell(ws, 'U', ao_row, -2598368, red_font)  # AO under-provided
    ws[f'U{ao_row}'].fill = disb_fill

    ws[f'V{ao_row}'].fill = source_fill
    num_cell(ws, 'V', ao_row, 6900669, red_font)  # AO adjusted: pays MORE

# Total row — add U and V
if total_sh_row:
    ws[f'U{total_sh_row}'].fill = total_fill
    num_cell(ws, 'U', total_sh_row, 0, bold_font)  # Net zero

    ws[f'V{total_sh_row}'].fill = total_fill
    num_cell(ws, 'V', total_sh_row, 8963126, bold_font)  # Same total

# Fix the explanation rows
if fund_row:
    exp_row = fund_row + 1
    # Unmerge if merged
    merges_to_remove = []
    for merge in ws.merged_cells.ranges:
        if merge.min_row == exp_row and merge.min_col == 11:
            merges_to_remove.append(merge)
    for m in merges_to_remove:
        ws.unmerge_cells(str(m))

    ws.merge_cells(f'K{exp_row}:V{exp_row}')
    ws[f'K{exp_row}'].value = '    → Oct-Dec: Fund provided EUR 5.74M (cashed guarantee) against EUR 3.04M required. Over-provided EUR 2.60M. Adjusted Apr-Jun: EUR 2.06M (reduced by disbalance credit).'
    ws[f'K{exp_row}'].font = small_font

if ao_row:
    exp_row = ao_row + 1
    merges_to_remove = []
    for merge in ws.merged_cells.ranges:
        if merge.min_row == exp_row and merge.min_col == 11:
            merges_to_remove.append(merge)
    for m in merges_to_remove:
        ws.unmerge_cells(str(m))

    ws.merge_cells(f'K{exp_row}:V{exp_row}')
    ws[f'K{exp_row}'].value = '    → Oct-Dec: EUR 0 provided (EUR 2.81M shortfall). Adjusted Apr-Jun: EUR 6.90M = 48% share (EUR 4.30M) + disbalance catch-up (EUR 2.60M). EUR 2.5M (Mar) accounted separately.'
    ws[f'K{exp_row}'].font = small_font

# Now fix the SUMMARY section at the bottom — find it and update AO figure
for r in range(total_sh_row + 1, 150):
    val = ws.cell(row=r, column=11).value
    if val and 'AO (48%)' in str(val) and 'Apr-Jun' in str(val):
        num_cell(ws, 'S', r, 6900669, bold_font)
        print(f"Updated summary AO row {r} to 6,900,669")
        break
    elif val and 'Fund (52%)' in str(val) and 'Apr-Jun' in str(val):
        num_cell(ws, 'S', r, 2062458, bold_font)
        print(f"Updated summary Fund row {r} to 2,062,458")

# Find and update the total to provide and verification rows
for r in range(total_sh_row + 1, 150):
    val = ws.cell(row=r, column=11).value
    if val and 'Total to be provided' in str(val):
        new_total = 2062458 + 6900669 + 3120000
        num_cell(ws, 'S', r, new_total, green_bold)
        print(f"Updated total to provide row {r} to {new_total:,}")
    elif val and 'Verification' in str(val):
        total_provided = 15051718
        total_to_provide = 2062458 + 6900669 + 3120000
        balance = -27134844 + total_provided + total_to_provide
        # Unmerge
        merges_to_remove = []
        for merge in ws.merged_cells.ranges:
            if merge.min_row == r and merge.min_col == 11:
                merges_to_remove.append(merge)
        for m in merges_to_remove:
            ws.unmerge_cells(str(m))
        ws.merge_cells(f'K{r}:V{r}')
        ws[f'K{r}'].value = f'Verification: -27,134,844 + {total_provided:,} (provided) + {total_to_provide:,} (to provide) = {balance:,}'
        ws[f'K{r}'].font = Font(name='Calibri', size=9, italic=True, color='006600' if abs(balance) < 100 else 'CC0000')
        print(f"Updated verification row {r}: balance = {balance}")
    elif val and 'Remaining Gap' in str(val):
        remaining = -27134844 + 15051718
        num_cell(ws, 'S', r, remaining, big_total_font)
        print(f"Updated remaining gap row {r} to {remaining:,}")

# Widen columns U and V
ws.column_dimensions['U'].width = 14
ws.column_dimensions['V'].width = 16

wb.save(dst)
print(f'\nSaved: {dst}')
