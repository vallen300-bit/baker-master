"""Baker AI — File-type text extractors.

Each extractor takes a file path and returns extracted text as a string.
Supported: .txt, .md, .pdf, .csv, .xlsx, .json, .jpg, .jpeg, .png, .heic, .webp
"""
import base64
import csv
import io
import json
import logging
import os
from pathlib import Path
from typing import Optional

import anthropic

logger = logging.getLogger("baker.ingest.extractors")

# Supported extensions mapped to their extractor functions
SUPPORTED_EXTENSIONS = {
    ".txt", ".md", ".pdf", ".csv", ".xlsx", ".json",
    ".jpg", ".jpeg", ".png", ".heic", ".webp",
}

# Image extensions handled by the image extractor
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".heic", ".webp"}


def extract(filepath: Path) -> str:
    """Extract text from a file based on its extension.

    Args:
        filepath: Path to the file to extract text from.

    Returns:
        Extracted text content as a string.

    Raises:
        ValueError: If file type is not supported.
        FileNotFoundError: If file doesn't exist.
    """
    if not filepath.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    ext = filepath.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file type '{ext}'. "
            f"Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    extractor = _EXTRACTORS[ext]
    text = extractor(filepath)

    if not text or not text.strip():
        logger.warning("Extracted empty text from %s", filepath.name)
        return ""

    return text.strip()


def _extract_text(filepath: Path) -> str:
    """Extract plain text / markdown files."""
    return filepath.read_text(encoding="utf-8", errors="replace")


def _extract_pdf(filepath: Path) -> str:
    """Extract text from PDF using pdfplumber."""
    try:
        import pdfplumber
    except ImportError:
        raise ImportError(
            "pdfplumber is required for PDF extraction. "
            "Install it: pip install pdfplumber"
        )

    pages = []
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
    return "\n\n".join(pages)


def _extract_csv(filepath: Path) -> str:
    """Extract CSV as row-per-line text with headers."""
    with open(filepath, "r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return ""

    headers = rows[0]
    lines = []
    for row in rows[1:]:
        pairs = []
        for h, v in zip(headers, row):
            if v.strip():
                pairs.append(f"{h}: {v}")
        if pairs:
            lines.append("; ".join(pairs))

    return "\n".join(lines)


def _extract_xlsx(filepath: Path) -> str:
    """Extract Excel spreadsheet as text (all sheets)."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel extraction. "
            "Install it: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h is not None else "" for h in rows[0]]
        lines = [f"[Sheet: {sheet_name}]"]

        for row in rows[1:]:
            pairs = []
            for h, v in zip(headers, row):
                if v is not None and str(v).strip():
                    pairs.append(f"{h}: {v}")
            if pairs:
                lines.append("; ".join(pairs))

        parts.append("\n".join(lines))

    wb.close()
    return "\n\n".join(parts)


def _extract_json(filepath: Path) -> str:
    """Extract JSON — handles Baker's {texts: [...]} format and generic JSON."""
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Baker-native format: {"texts": [{"text": "...", "metadata": {...}}, ...]}
    if isinstance(data, dict) and "texts" in data:
        items = data["texts"]
        parts = []
        for item in items:
            if isinstance(item, dict) and "text" in item:
                parts.append(item["text"])
        if parts:
            return "\n\n".join(parts)

    # Generic JSON — pretty-print
    return json.dumps(data, indent=2, ensure_ascii=False, default=str)


# -------------------------------------------------------
# Image extraction via Claude Vision
# -------------------------------------------------------

_CARD_PROMPT = """You are a precise data extractor. This image is a business card.

Extract ALL visible information and return ONLY valid JSON (no markdown, no commentary):

{
  "name": "Full Name",
  "company": "Company / Organization",
  "role": "Title / Position",
  "email": "email@example.com",
  "phone": "+1234567890",
  "address": "Full address if visible",
  "website": "URL if visible",
  "notes": "Any other details (LinkedIn, social handles, etc.)"
}

Rules:
- Use null for fields not visible on the card.
- For phone numbers, include country code if shown.
- If multiple emails/phones, pick the primary one.
- Return raw JSON only — no ```json``` fences, no explanation."""

_WHITEBOARD_PROMPT = """You are a meticulous note-taker. This image shows a whiteboard, notepad, or handwritten document.

1. First, transcribe ALL visible text exactly as written (preserve structure, headings, bullets, arrows).
2. Then, synthesize a clean summary of the key points, decisions, and action items.

Format your response as:

## Transcription
[Exact text as written on the board/page]

## Summary
[Clean synthesis of key points]

## Action Items
- [Any action items or tasks identified]

Rules:
- Preserve the original structure (columns, boxes, arrows → describe spatial layout).
- If text is partially illegible, use [illegible] markers.
- Include all diagrams/drawings as text descriptions."""

_AUTO_CLASSIFY_PROMPT = """Look at this image and classify it as exactly one of:
- "card" — a business card with contact information
- "whiteboard" — a whiteboard, notepad, handwritten note, or document photo

Reply with ONLY the single word: card OR whiteboard"""

# MIME type mapping for Claude Vision API
_IMAGE_MIME = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
    ".heic": "image/jpeg",  # Converted to JPEG before sending
}


def _convert_heic_to_jpeg(filepath: Path) -> bytes:
    """Convert HEIC image to JPEG bytes using pillow-heif."""
    try:
        from pillow_heif import register_heif_opener
        from PIL import Image
    except ImportError:
        raise ImportError(
            "pillow-heif and Pillow are required for HEIC conversion. "
            "Install: pip install pillow-heif Pillow"
        )

    register_heif_opener()
    img = Image.open(filepath)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=90)
    return buf.getvalue()


def _load_image_bytes(filepath: Path) -> tuple[bytes, str]:
    """Load image bytes, converting HEIC if needed. Returns (bytes, mime_type)."""
    ext = filepath.suffix.lower()

    if ext == ".heic":
        img_bytes = _convert_heic_to_jpeg(filepath)
        mime = "image/jpeg"
    else:
        img_bytes = filepath.read_bytes()
        mime = _IMAGE_MIME.get(ext, "image/jpeg")

    return img_bytes, mime


def _call_claude_vision(img_bytes: bytes, mime: str, prompt: str, max_tokens: int = 2000) -> str:
    """Send image to Claude Sonnet via vision API and return text response."""
    client = anthropic.Anthropic()  # Uses ANTHROPIC_API_KEY env var
    b64 = base64.standard_b64encode(img_bytes).decode("utf-8")

    response = client.messages.create(
        model="claude-sonnet-4-5-20250929",
        max_tokens=max_tokens,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": mime,
                            "data": b64,
                        },
                    },
                    {
                        "type": "text",
                        "text": prompt,
                    },
                ],
            }
        ],
    )
    return response.content[0].text


def _classify_image(img_bytes: bytes, mime: str) -> str:
    """Auto-classify image as 'card' or 'whiteboard'."""
    result = _call_claude_vision(img_bytes, mime, _AUTO_CLASSIFY_PROMPT, max_tokens=10)
    classification = result.strip().lower()
    if classification in ("card", "whiteboard"):
        return classification
    # Default to whiteboard if unclear
    logger.warning("Image classification unclear ('%s'), defaulting to whiteboard", classification)
    return "whiteboard"


def extract_image(filepath: Path, image_type: Optional[str] = None) -> str:
    """Extract text/data from an image using Claude Vision.

    Args:
        filepath: Path to image file.
        image_type: 'card', 'whiteboard', or None for auto-detect.

    Returns:
        Extracted text. For cards, returns JSON string with contact fields.
        For whiteboards, returns structured transcription + summary.
    """
    img_bytes, mime = _load_image_bytes(filepath)

    # Auto-classify if type not specified
    if not image_type or image_type == "auto":
        image_type = _classify_image(img_bytes, mime)
        logger.info("Auto-classified %s as: %s", filepath.name, image_type)

    if image_type == "card":
        prompt = _CARD_PROMPT
    else:
        prompt = _WHITEBOARD_PROMPT

    text = _call_claude_vision(img_bytes, mime, prompt)
    return text


def _extract_image_default(filepath: Path) -> str:
    """Default image extractor (auto-detect mode) for the _EXTRACTORS dispatch."""
    return extract_image(filepath, image_type=None)


def parse_card_json(text: str) -> Optional[dict]:
    """Parse the JSON output from a business card extraction.

    Returns dict with contact fields, or None if parsing fails.
    """
    try:
        # Strip any accidental markdown fences
        cleaned = text.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("\n", 1)[1]
        if cleaned.endswith("```"):
            cleaned = cleaned.rsplit("```", 1)[0]
        cleaned = cleaned.strip()

        data = json.loads(cleaned)
        return data
    except (json.JSONDecodeError, IndexError) as e:
        logger.warning("Failed to parse card JSON: %s", e)
        return None


# Extension → extractor mapping
_EXTRACTORS = {
    ".txt": _extract_text,
    ".md": _extract_text,
    ".pdf": _extract_pdf,
    ".csv": _extract_csv,
    ".xlsx": _extract_xlsx,
    ".json": _extract_json,
    ".jpg": _extract_image_default,
    ".jpeg": _extract_image_default,
    ".png": _extract_image_default,
    ".heic": _extract_image_default,
    ".webp": _extract_image_default,
}
